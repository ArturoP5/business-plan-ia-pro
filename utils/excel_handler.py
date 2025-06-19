import pandas as pd
import streamlit as st
from io import BytesIO
import xlsxwriter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

# ESTILOS CORPORATIVOS
COLOR_PRIMARIO = "1E3A8A"  # Azul oscuro
COLOR_SECUNDARIO = "3B82F6"  # Azul medio
COLOR_FONDO = "F3F4F6"  # Gris claro
COLOR_EXITO = "10B981"  # Verde
COLOR_ERROR = "EF4444"  # Rojo

# Estilos predefinidos
ESTILO_HEADER = {
    'font': Font(bold=True, color="FFFFFF", size=12),
    'fill': PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid"),
    'alignment': Alignment(horizontal="center", vertical="center"),
    'border': Border(
        left=Side(style='thin', color="FFFFFF"),
        right=Side(style='thin', color="FFFFFF"),
        top=Side(style='thin', color="FFFFFF"),
        bottom=Side(style='thin', color="FFFFFF")
    )
}

ESTILO_SUBTITULO = {
    'font': Font(bold=True, size=11),
    'fill': PatternFill(start_color=COLOR_SECUNDARIO, end_color=COLOR_SECUNDARIO, fill_type="solid"),
    'alignment': Alignment(horizontal="left", vertical="center")
}

ESTILO_TOTAL = {
    'font': Font(bold=True, size=11),
    'fill': PatternFill(start_color=COLOR_FONDO, end_color=COLOR_FONDO, fill_type="solid"),
    'border': Border(
        top=Side(style='double'),
        bottom=Side(style='double')
    )
}

ESTILO_CELDA = {
    'border': Border(
        left=Side(style='thin', color="D1D5DB"),
        right=Side(style='thin', color="D1D5DB"),
        top=Side(style='thin', color="D1D5DB"),
        bottom=Side(style='thin', color="D1D5DB")
    )
}

# Bordes específicos
BORDE_EXTERIOR = Border(
    left=Side(style='medium', color="000000"),
    right=Side(style='medium', color="000000"),
    top=Side(style='medium', color="000000"),
    bottom=Side(style='medium', color="000000")
)

BORDE_INTERIOR_COLUMNA = Border(
    left=Side(style='thin', color="D1D5DB"),
    right=Side(style='thin', color="D1D5DB"),
    top=Side(style='dotted', color="D1D5DB"),
    bottom=Side(style='dotted', color="D1D5DB")
)

BORDE_HEADER_DATOS = Border(
    left=Side(style='thin', color="D1D5DB"),
    right=Side(style='thin', color="D1D5DB"),
    top=Side(style='thin', color="000000"),
    bottom=Side(style='thin', color="000000")
)

def crear_plantilla_excel(sector="General"):
    """
    Crea una plantilla Excel con la estructura exacta de la app
    """
    output = BytesIO()

    # Crear workbook con openpyxl para diseño avanzado
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    wb.remove(wb.active)
    
    # DEBUG: Crear solo ciertas hojas
    HOJAS_A_CREAR = ['INSTRUCCIONES', 'INFO_GENERAL', 'PYL']  # Añadir o quitar según necesites
    
    # Función helper para aplicar estilos
    def aplicar_estilo(celda, estilo_dict):
        for attr, value in estilo_dict.items():
            setattr(celda, attr, value)
    
    # Función para ajustar ancho de columnas
    def ajustar_columnas(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Función para aplicar bordes a un rango
    def aplicar_bordes_tabla(ws, fila_inicio, col_inicio, fila_fin, col_fin, es_primera_fila_datos=False):
        # Bordes exteriores
        for row in range(fila_inicio, fila_fin + 1):
            for col in range(col_inicio, col_fin + 1):
                celda = ws.cell(row=row, column=col)
                
                # Determinar qué tipo de borde aplicar
                if row == fila_inicio and es_primera_fila_datos:
                    celda.border = BORDE_HEADER_DATOS
                else:
                    celda.border = BORDE_INTERIOR_COLUMNA
                
                # Bordes exteriores más gruesos
                if row == fila_inicio:
                    celda.border = Border(
                        left=celda.border.left if col > col_inicio else Side(style='medium'),
                        right=celda.border.right if col < col_fin else Side(style='medium'),
                        top=Side(style='medium'),
                        bottom=celda.border.bottom
                    )
                elif row == fila_fin:
                    celda.border = Border(
                        left=celda.border.left if col > col_inicio else Side(style='medium'),
                        right=celda.border.right if col < col_fin else Side(style='medium'),
                        top=celda.border.top,
                        bottom=Side(style='medium')
                    )
                
                if col == col_inicio:
                    celda.border = Border(
                        left=Side(style='medium'),
                        right=celda.border.right,
                        top=celda.border.top,
                        bottom=celda.border.bottom
                    )
                elif col == col_fin:
                    celda.border = Border(
                        left=celda.border.left,
                        right=Side(style='medium'),
                        top=celda.border.top,
                        bottom=celda.border.bottom
                    )

    # ==== HOJA 1: INSTRUCCIONES ====
    ws_inst = wb.create_sheet("📚 INSTRUCCIONES")
    ws_inst.sheet_properties.tabColor = COLOR_PRIMARIO
    
    # Título principal
    ws_inst.merge_cells('A1:F1')
    titulo = ws_inst['A1']
    titulo.value = "PLANTILLA BUSINESS PLAN PROFESIONAL"
    aplicar_estilo(titulo, ESTILO_HEADER)
    ws_inst.row_dimensions[1].height = 30
    
    # Subtítulo
    ws_inst.merge_cells('A3:F3')
    subtitulo = ws_inst['A3']
    subtitulo.value = f"Plantilla optimizada para sector: {sector}"
    subtitulo.font = Font(italic=True, size=11)
    subtitulo.alignment = Alignment(horizontal="center")
    
    # Instrucciones
    instrucciones = [
        ("A5", "📋 CÓMO USAR ESTA PLANTILLA:", ESTILO_SUBTITULO),
        ("A7", "1. Complete la información en orden: Info General → PYL → Balance → Laborales"),
        ("A8", "2. Los campos en GRIS están bloqueados (se calculan automáticamente)"),
        ("A9", "3. Use las listas desplegables donde estén disponibles"),
        ("A10", "4. Revise la hoja ✅ VERIFICACIÓN para comprobar coherencia"),
        ("A12", "💡 TIPS:", ESTILO_SUBTITULO),
        ("A14", "• Las celdas en AMARILLO requieren su atención especial"),
        ("A15", "• Los totales se calculan automáticamente"),
        ("A16", "• Guarde frecuentemente su trabajo"),
        ("A17", f"• Valores típicos del sector {sector} ya están pre-cargados")
    ]
    
    for cell, text, *style in instrucciones:
        ws_inst[cell].value = text
        if style:
            aplicar_estilo(ws_inst[cell], style[0])
    
    # Ajustar anchos
    ws_inst.column_dimensions['A'].width = 80
    
    # Agregar logo/espacio para branding
    ws_inst.merge_cells('H1:J4')
    logo_space = ws_inst['H1']
    logo_space.value = "[LOGO]"
    logo_space.alignment = Alignment(horizontal="center", vertical="center")
    logo_space.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Valores por defecto según sector
    valores_sector = {
        "General": {
            "ventas_1": 1000000, "ventas_2": 1100000, "ventas_3": 1200000,
            "costos_var": 40, "gastos_personal": 200000, "gastos_generales": 50000,
            "gastos_marketing": 20000, "margen_esperado": "15-20%"
        },
        "Hostelería": {
            "ventas_1": 800000, "ventas_2": 900000, "ventas_3": 1000000,
            "costos_var": 45, "gastos_personal": 280000, "gastos_generales": 60000,
            "gastos_marketing": 15000, "margen_esperado": "10-15%"
        },
        "Tecnología": {
            "ventas_1": 2000000, "ventas_2": 3000000, "ventas_3": 4500000,
            "costos_var": 25, "gastos_personal": 800000, "gastos_generales": 120000,
            "gastos_marketing": 200000, "margen_esperado": "25-35%"
        },
        "Ecommerce": {
            "ventas_1": 1500000, "ventas_2": 2000000, "ventas_3": 2800000,
            "costos_var": 65, "gastos_personal": 150000, "gastos_generales": 80000,
            "gastos_marketing": 100000, "margen_esperado": "8-12%"
        },
        "Industrial": {
            "ventas_1": 5000000, "ventas_2": 5500000, "ventas_3": 6000000,
            "costos_var": 60, "gastos_personal": 500000, "gastos_generales": 150000,
            "gastos_marketing": 50000, "margen_esperado": "12-18%"
        }
    }
    
    # Obtener valores del sector seleccionado
    valores = valores_sector.get(sector, valores_sector["General"])
    
    # Definir año actual para usar en toda la plantilla
    año_actual = datetime.now().year
    
    # Crear el libro de Excel
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#1E40AF',
            'font_color': 'white',
            'border': 1
        })
        
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'bg_color': '#E0E7FF'
        })
        
    # ==== HOJA 2: INFORMACIÓN GENERAL ====
    # Crear hoja Info General
    ws_info = wb.create_sheet("Info General")
    ws_info.sheet_properties.tabColor = COLOR_SECUNDARIO
    
    # Título
    ws_info.merge_cells('A1:C1')
    titulo_info = ws_info['A1']
    titulo_info.value = "INFORMACIÓN GENERAL DE LA EMPRESA"
    aplicar_estilo(titulo_info, ESTILO_HEADER)
    ws_info.row_dimensions[1].height = 25
    
    # Headers de columnas
    headers = ['Campo', 'Valor', 'Instrucciones']
    for col, header in enumerate(headers, 1):
        celda = ws_info.cell(row=3, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    
    # Ancho de columnas
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 30
    ws_info.column_dimensions['C'].width = 40
    
    # Datos con formato
    campos_info = [
        ('Nombre de la empresa', '', 'Razón social completa'),
        ('Sector', sector, 'Seleccione de la lista'),
        ('País', 'España', 'País de operaciones'),
        ('Año fundación', '', 'Año de constitución (YYYY)'),
        ('Empresa familiar', 'No', 'Sí/No'),
        ('Cuentas auditadas', 'Sí', 'Sí/No'),
        ('Moneda', 'EUR', 'EUR/USD/GBP')
    ]
    
    for idx, (campo, valor, instruccion) in enumerate(campos_info, 4):
        # Campo
        celda_campo = ws_info.cell(row=idx, column=1)
        celda_campo.value = campo
        aplicar_estilo(celda_campo, ESTILO_CELDA)
        
        # Valor
        celda_valor = ws_info.cell(row=idx, column=2)
        celda_valor.value = valor
        aplicar_estilo(celda_valor, ESTILO_CELDA)
        celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        
        # Instrucción
        celda_inst = ws_info.cell(row=idx, column=3)
        celda_inst.value = instruccion
        aplicar_estilo(celda_inst, ESTILO_CELDA)
        celda_inst.font = Font(italic=True, size=10, color="6B7280")

        # Aplicar bordes a Info General
    aplicar_bordes_tabla(ws_info, 3, 1, 10, 3)
    
    

    df_info = pd.DataFrame({
        'Campo': [
            'Nombre de la empresa',
            'Sector',
            'País', 
            'Año de Fundación',
            '¿Empresa familiar?',
            '¿Cuentas auditadas?',
            'Moneda'
        ],
        'Valor': [''] * 7,
        'Opciones': [
            'Texto libre',
            'Hostelería/Tecnología/Ecommerce/Consultoría/Retail/Servicios/Automoción/Industrial/Otro',
            'España/Francia/Alemania/Reino Unido/Estados Unidos/Portugal/Italia/Países Bajos/Irlanda/Bélgica',
            'Número (ej: 2015)',
            'Sí/No',
            'Sí/No',
            'EUR/USD/GBP/CHF'
        ]
    })
    #df_info.to_excel(writer, sheet_name='Información General', index=False)
        
    # HOJA 2: DATOS HISTÓRICOS P&L

    # ==== HOJA 3: DATOS HISTÓRICOS PYL ====
    ws_pyl = wb.create_sheet("Datos Históricos PYL")
    ws_pyl.sheet_properties.tabColor = COLOR_SECUNDARIO
    
    # Título
    ws_pyl.merge_cells('A1:E1')
    titulo_pyl = ws_pyl['A1']
    titulo_pyl.value = "CUENTA DE RESULTADOS HISTÓRICA"
    aplicar_estilo(titulo_pyl, ESTILO_HEADER)
    ws_pyl.row_dimensions[1].height = 25
    
    # Preparar años
    año_actual = datetime.now().year
    df_pyl = pd.DataFrame({
        'Concepto': [
            'Ventas',
            'Costos Variables (%)',
            'Gastos de Personal',
            'Gastos Generales',
            'Gastos de Marketing'
        ],
        f'Año {año_actual-3}': [valores['ventas_1'], valores['costos_var'], valores['gastos_personal'], valores['gastos_generales'], valores['gastos_marketing']],
        f'Año {año_actual-2}': [valores['ventas_2'], valores['costos_var'], valores['gastos_personal'], valores['gastos_generales'], valores['gastos_marketing']],
        f'Año {año_actual-1}': [valores['ventas_3'], valores['costos_var'], valores['gastos_personal'], valores['gastos_generales'], valores['gastos_marketing']]
    })
    df_pyl.to_excel(writer, sheet_name='Datos Históricos PYL', index=False)

    # Aplicar formato a PYL en openpyxl
    headers_pyl = ['Concepto', f'Año {año_actual-3}', f'Año {año_actual-2}', f'Año {año_actual-1}']
    for col, header in enumerate(headers_pyl, 1):
        celda = ws_pyl.cell(row=3, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    
    # Datos del PYL con formato
    conceptos_pyl = [
        ('Ventas', valores['ventas_1'], valores['ventas_2'], valores['ventas_3']),
        ('Costos Variables (%)', 40, 40, 40),
        ('Gastos de Personal', valores['gastos_personal'], valores['gastos_personal'], valores['gastos_personal']),
        ('Gastos Generales', valores['gastos_generales'], valores['gastos_generales'], valores['gastos_generales']),
        ('Gastos de Marketing', valores['gastos_marketing'], valores['gastos_marketing'], valores['gastos_marketing'])
    ]
    
    for idx, datos in enumerate(conceptos_pyl, 4):
        for col, valor in enumerate(datos, 1):
            celda = ws_pyl.cell(row=idx, column=col)
            celda.value = valor
            aplicar_estilo(celda, ESTILO_CELDA)
            if col > 1:  # Columnas numéricas
                celda.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
                if "%" in str(datos[0]):
                    celda.value = valor / 100  # Convertir a decimal
                    celda.number_format = '0.00%'
                else:
                    celda.number_format = '#,##0'
    
    # Ancho de columnas
    ws_pyl.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws_pyl.column_dimensions[col].width = 15
    # Aplicar bordes a la tabla PYL
    #aplicar_bordes_tabla(ws_pyl, 3, 1, 8, 4, es_primera_fila_datos=True)
        
    # HOJA 3: BALANCE - ACTIVO
    # ==== HOJA 4: BALANCE - ACTIVO ====
    ws_activo = wb.create_sheet("Balance - Activo")
    ws_activo.sheet_properties.tabColor = COLOR_SECUNDARIO
    
    # Título
    ws_activo.merge_cells('A1:C1')
    titulo_activo = ws_activo['A1']
    titulo_activo.value = "BALANCE - ACTIVO"
    aplicar_estilo(titulo_activo, ESTILO_HEADER)
    ws_activo.row_dimensions[1].height = 25
    
    # Headers
    headers_balance = ['Concepto', f'Año {año_actual-1}', 'Notas']
    for col, header in enumerate(headers_balance, 1):
        celda = ws_activo.cell(row=3, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    
    # Ancho de columnas
    ws_activo.column_dimensions['A'].width = 35
    ws_activo.column_dimensions['B'].width = 18
    ws_activo.column_dimensions['C'].width = 30

    # Crear lista de conceptos
    conceptos_activo = [
        'Caja y bancos',
        'Inversiones financieras temporales',
        'Clientes comerciales',
        'Otros deudores',
        'Administraciones públicas deudoras',
        'Inventarios',
        'Gastos anticipados',
        'Activos por impuesto diferido CP',
        '',
        'ACTIVO NO CORRIENTE',
        'Inmovilizado material bruto',
        'Amortización acumulada material',
        'Activos intangibles brutos',
        'Amortización acumulada intangibles',
        'Participaciones en empresas',
        'Créditos a largo plazo',
        'Fianzas y depósitos',
        'Activos por impuesto diferido LP'
    ]
    
    df_activo = pd.DataFrame({
        'ACTIVO CORRIENTE': conceptos_activo,
        f'Año {año_actual-1}': [0] * len(conceptos_activo)
    })
    df_activo.to_excel(writer, sheet_name='Balance - Activo', index=False)

    # Datos del Balance Activo con formato
    fila = 4
    # ACTIVO CORRIENTE - Título
    celda_titulo = ws_activo.cell(row=fila, column=1)
    celda_titulo.value = "ACTIVO CORRIENTE"
    celda_titulo.font = Font(bold=True, size=11, color="FFFFFF")
    celda_titulo.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    ws_activo.merge_cells(f'A{fila}:C{fila}')
    fila += 1
    
    # Conceptos de Activo Corriente
    activo_corriente = [
        ('Caja y bancos', 0, 'Efectivo disponible'),
        ('Inversiones financieras temporales', 0, 'Inversiones < 1 año'),
        ('Clientes comerciales', 0, 'Facturas pendientes cobro'),
        ('Otros deudores', 0, 'Otros créditos CP'),
        ('Administraciones públicas deudoras', 0, 'IVA a compensar, etc.'),
        ('Inventarios', 0, 'Stock de productos'),
        ('Gastos anticipados', 0, 'Pagos por adelantado'),
        ('Activos por impuesto diferido CP', 0, 'Créditos fiscales CP')
    ]
    
    for concepto, valor, nota in activo_corriente:
        ws_activo.cell(row=fila, column=1).value = concepto
        celda_valor = ws_activo.cell(row=fila, column=2)
        celda_valor.value = valor
        celda_valor.number_format = '#,##0'
        celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        ws_activo.cell(row=fila, column=3).value = nota
        ws_activo.cell(row=fila, column=3).font = Font(italic=True, size=9, color="6B7280")
        fila += 1
    
    # Total Activo Corriente
    celda_total_ac = ws_activo.cell(row=fila, column=1)
    celda_total_ac.value = "TOTAL ACTIVO CORRIENTE"
    aplicar_estilo(celda_total_ac, ESTILO_TOTAL)
    celda_formula_ac = ws_activo.cell(row=fila, column=2)
    celda_formula_ac.value = "=SUM(B5:B12)"
    celda_formula_ac.number_format = '#,##0'
    aplicar_estilo(celda_formula_ac, ESTILO_TOTAL)
    celda_formula_ac.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    fila += 2
    
    # ACTIVO NO CORRIENTE - Título
    celda_titulo_nc = ws_activo.cell(row=fila, column=1)
    celda_titulo_nc.value = "ACTIVO NO CORRIENTE"
    celda_titulo_nc.font = Font(bold=True, size=11, color="FFFFFF")
    celda_titulo_nc.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    ws_activo.merge_cells(f'A{fila}:C{fila}')
    fila += 1
    
    # Aplicar bordes
    #aplicar_bordes_tabla(ws_activo, 3, 1, fila-1, 3)

    # Conceptos de Activo No Corriente
    activo_no_corriente = [
        ('Inmovilizado material bruto', 0, 'Valor de compra activos'),
        ('Amortización acumulada material', 0, 'Desgaste acumulado (-)'),
        ('Activos intangibles brutos', 0, 'Software, patentes, etc.'),
        ('Amortización acumulada intangibles', 0, 'Amortización intangibles (-)'),
        ('Participaciones en empresas', 0, 'Inversiones en otras empresas'),
        ('Créditos a largo plazo', 0, 'Préstamos concedidos LP'),
        ('Fianzas y depósitos', 0, 'Garantías constituidas'),
        ('Activos por impuesto diferido LP', 0, 'Créditos fiscales LP')
    ]
    
    for concepto, valor, nota in activo_no_corriente:
        ws_activo.cell(row=fila, column=1).value = concepto
        celda_valor = ws_activo.cell(row=fila, column=2)
        celda_valor.value = valor
        celda_valor.number_format = '#,##0'
        celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        ws_activo.cell(row=fila, column=3).value = nota
        ws_activo.cell(row=fila, column=3).font = Font(italic=True, size=9, color="6B7280")
        fila += 1
    
    # Total Activo No Corriente
    celda_total_anc = ws_activo.cell(row=fila, column=1)
    celda_total_anc.value = "TOTAL ACTIVO NO CORRIENTE"
    aplicar_estilo(celda_total_anc, ESTILO_TOTAL)
    celda_formula_anc = ws_activo.cell(row=fila, column=2)
    celda_formula_anc.value = f"=SUM(B{fila-8}:B{fila-1})"

    celda_formula_anc.number_format = '#,##0'
    aplicar_estilo(celda_formula_anc, ESTILO_TOTAL)
    celda_formula_anc.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    fila += 2
    
    # TOTAL ACTIVO
    celda_total_activo = ws_activo.cell(row=fila, column=1)
    celda_total_activo.value = "TOTAL ACTIVO"
    celda_total_activo.font = Font(bold=True, size=12, color="FFFFFF")
    celda_total_activo.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    
    celda_formula_total = ws_activo.cell(row=fila, column=2)
    celda_formula_total.value = "=B13+B24"# Suma de totales corriente y no corriente
    celda_formula_total.number_format = '#,##0'
    celda_formula_total.font = Font(bold=True, size=12, color="FFFFFF")
    celda_formula_total.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    
    # Aplicar bordes finales
    #aplicar_bordes_tabla(ws_activo, 15, 1, fila, 3)
   

    # HOJA 4: BALANCE - PASIVO
    # ==== HOJA 5: BALANCE - PASIVO ====
    ws_pasivo = wb.create_sheet("Balance - Pasivo")
    ws_pasivo.sheet_properties.tabColor = COLOR_SECUNDARIO
    
    # Título
    ws_pasivo.merge_cells('A1:C1')
    titulo_pasivo = ws_pasivo['A1']
    titulo_pasivo.value = "BALANCE - PASIVO"
    aplicar_estilo(titulo_pasivo, ESTILO_HEADER)
    ws_pasivo.row_dimensions[1].height = 25
    
    # Headers
    for col, header in enumerate(headers_balance, 1):
        celda = ws_pasivo.cell(row=3, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    
    # Ancho de columnas
    ws_pasivo.column_dimensions['A'].width = 35
    ws_pasivo.column_dimensions['B'].width = 18
    ws_pasivo.column_dimensions['C'].width = 30
    
    # Datos del Balance Pasivo
    fila = 4
    # PASIVO CORRIENTE - Título
    celda_titulo = ws_pasivo.cell(row=fila, column=1)
    celda_titulo.value = "PASIVO CORRIENTE"
    celda_titulo.font = Font(bold=True, size=11, color="FFFFFF")
    celda_titulo.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    ws_pasivo.merge_cells(f'A{fila}:C{fila}')
    fila += 1
    
    # Conceptos de Pasivo Corriente
    pasivo_corriente = [
        ('Proveedores comerciales', 0, 'Facturas pendientes pago'),
        ('Acreedores por servicios', 0, 'Servicios pendientes pago'),
        ('Anticipos de clientes', 0, 'Cobros anticipados'),
        ('Remuneraciones pendientes', 0, 'Nóminas pendientes'),
        ('Administraciones públicas acreedoras', 0, 'IRPF, IVA, SS pendientes'),
        ('Provisiones a corto plazo', 0, 'Provisiones < 1 año'),
        ('Deuda bancaria corto plazo', 0, 'Préstamos < 1 año'),
        ('Otros pasivos corrientes', 0, 'Otras deudas CP')
    ]
    
    for concepto, valor, nota in pasivo_corriente:
        ws_pasivo.cell(row=fila, column=1).value = concepto
        celda_valor = ws_pasivo.cell(row=fila, column=2)
        celda_valor.value = valor
        celda_valor.number_format = '#,##0'
        celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        ws_pasivo.cell(row=fila, column=3).value = nota
        ws_pasivo.cell(row=fila, column=3).font = Font(italic=True, size=9, color="6B7280")
        fila += 1
    
    # Total Pasivo Corriente
    celda_total_pc = ws_pasivo.cell(row=fila, column=1)
    celda_total_pc.value = "TOTAL PASIVO CORRIENTE"
    aplicar_estilo(celda_total_pc, ESTILO_TOTAL)
    celda_formula_pc = ws_pasivo.cell(row=fila, column=2)
    celda_formula_pc.value = "=SUM(B5:B12)"
    celda_formula_pc.number_format = '#,##0'
    aplicar_estilo(celda_formula_pc, ESTILO_TOTAL)
    celda_formula_pc.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    fila += 2
    df_pasivo = pd.DataFrame({
        'PASIVO CORRIENTE': [
            'Proveedores comerciales',
            'Acreedores por servicios',
            'Anticipos de clientes',
            'Remuneraciones pendientes',
            'Administraciones públicas acreedoras',
            'Provisiones a corto plazo',
            'Otros pasivos corrientes',
            '',
            'PASIVO NO CORRIENTE',
            'Préstamos bancarios LP',
            'Hipoteca importe original',
            'Hipoteca meses transcurridos',
            'Leasing pendiente',
            'Otros préstamos LP',
            'Provisiones para riesgos',
            'Otras provisiones LP',
            'Pasivos por impuesto diferido'
        ],
        f'Año {año_actual-1}': [0] * 17
    })
    df_pasivo.to_excel(writer, sheet_name='Balance - Pasivo', index=False)

    # PASIVO NO CORRIENTE - Título
    celda_titulo_pnc = ws_pasivo.cell(row=fila, column=1)
    celda_titulo_pnc.value = "PASIVO NO CORRIENTE"
    celda_titulo_pnc.font = Font(bold=True, size=11, color="FFFFFF")
    celda_titulo_pnc.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    ws_pasivo.merge_cells(f'A{fila}:C{fila}')
    fila += 1
    
    # Conceptos de Pasivo No Corriente
    pasivo_no_corriente = [
        ('Préstamos bancarios LP', 0, 'Préstamos > 1 año'),
        ('Hipoteca pendiente', 0, 'Importe pendiente hipoteca'),
        ('Leasing pendiente', 0, 'Cuotas leasing pendientes'),
        ('Otros préstamos LP', 0, 'Otros préstamos > 1 año'),
        ('Provisiones para riesgos', 0, 'Provisiones largo plazo'),
        ('Otras provisiones LP', 0, 'Otras provisiones LP'),
        ('Pasivos por impuesto diferido', 0, 'Pasivos fiscales diferidos')
    ]
    
    for concepto, valor, nota in pasivo_no_corriente:
        ws_pasivo.cell(row=fila, column=1).value = concepto
        celda_valor = ws_pasivo.cell(row=fila, column=2)
        celda_valor.value = valor
        celda_valor.number_format = '#,##0'
        celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        ws_pasivo.cell(row=fila, column=3).value = nota
        ws_pasivo.cell(row=fila, column=3).font = Font(italic=True, size=9, color="6B7280")
        fila += 1
    
    # Total Pasivo No Corriente
    celda_total_pnc = ws_pasivo.cell(row=fila, column=1)
    celda_total_pnc.value = "TOTAL PASIVO NO CORRIENTE"
    aplicar_estilo(celda_total_pnc, ESTILO_TOTAL)
    celda_formula_pnc = ws_pasivo.cell(row=fila, column=2)
    celda_formula_pnc.value = f"=SUM(B{fila-7}:B{fila-1})"
    celda_formula_pnc.number_format = '#,##0'
    aplicar_estilo(celda_formula_pnc, ESTILO_TOTAL)
    celda_formula_pnc.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    fila += 2
    
    # TOTAL PASIVO
    celda_total_pasivo = ws_pasivo.cell(row=fila, column=1)
    celda_total_pasivo.value = "TOTAL PASIVO"
    celda_total_pasivo.font = Font(bold=True, size=12, color="FFFFFF")
    celda_total_pasivo.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    
    celda_formula_total = ws_pasivo.cell(row=fila, column=2)
    celda_formula_total.value = "=B13+B23"
    celda_formula_total.number_format = '#,##0'
    celda_formula_total.font = Font(bold=True, size=12, color="FFFFFF")
    celda_formula_total.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    
    # Aplicar bordes
    #aplicar_bordes_tabla(ws_pasivo, 3, 1, fila, 3)
    
    # HOJA 5: BALANCE - PATRIMONIO
    # ==== HOJA 6: BALANCE - PATRIMONIO ====
    ws_patrimonio = wb.create_sheet("Balance - Patrimonio")
    ws_patrimonio.sheet_properties.tabColor = COLOR_SECUNDARIO
    
    # Título
    ws_patrimonio.merge_cells('A1:C1')
    titulo_patrimonio = ws_patrimonio['A1']
    titulo_patrimonio.value = "BALANCE - PATRIMONIO NETO"
    aplicar_estilo(titulo_patrimonio, ESTILO_HEADER)
    ws_patrimonio.row_dimensions[1].height = 25
    
    # Headers
    for col, header in enumerate(headers_balance, 1):
        celda = ws_patrimonio.cell(row=3, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    
    # Ancho de columnas
    ws_patrimonio.column_dimensions['A'].width = 35
    ws_patrimonio.column_dimensions['B'].width = 18
    ws_patrimonio.column_dimensions['C'].width = 30
    
    # Datos del Patrimonio
    fila = 4
    # PATRIMONIO NETO - Título
    celda_titulo = ws_patrimonio.cell(row=fila, column=1)
    celda_titulo.value = "PATRIMONIO NETO"
    celda_titulo.font = Font(bold=True, size=11, color="FFFFFF")
    celda_titulo.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    ws_patrimonio.merge_cells(f'A{fila}:C{fila}')
    fila += 1
    
    # Conceptos de Patrimonio
    patrimonio_conceptos = [
        ('Capital social', 100000, 'Capital aportado socios'),
        ('Prima de emisión', 0, 'Sobreprecio acciones'),
        ('Reserva legal', 20000, '20% del capital social'),
        ('Otras reservas', 0, 'Reservas voluntarias'),
        ('Resultados ejercicios anteriores', 0, 'Beneficios/pérdidas acumuladas'),
        ('Resultado del ejercicio', 0, 'Beneficio/pérdida año actual'),
        ('Ajustes por cambio de valor', 0, 'Ajustes valoración'),
        ('Subvenciones de capital', 0, 'Subvenciones pendientes')
    ]
    
    for concepto, valor, nota in patrimonio_conceptos:
        ws_patrimonio.cell(row=fila, column=1).value = concepto
        celda_valor = ws_patrimonio.cell(row=fila, column=2)
        celda_valor.value = valor
        celda_valor.number_format = '#,##0'
        celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        ws_patrimonio.cell(row=fila, column=3).value = nota
        ws_patrimonio.cell(row=fila, column=3).font = Font(italic=True, size=9, color="6B7280")
        fila += 1
    
    # Total Patrimonio Neto
    celda_total_pn = ws_patrimonio.cell(row=fila, column=1)
    celda_total_pn.value = "TOTAL PATRIMONIO NETO"
    celda_total_pn.font = Font(bold=True, size=12, color="FFFFFF")
    celda_total_pn.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    
    celda_formula_pn = ws_patrimonio.cell(row=fila, column=2)
    celda_formula_pn.value = "=SUM(B5:B12)"
    celda_formula_pn.number_format = '#,##0'
    celda_formula_pn.font = Font(bold=True, size=12, color="FFFFFF")
    celda_formula_pn.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    fila += 2
    
    # VERIFICACIÓN BALANCE
    ws_patrimonio.merge_cells(f'A{fila}:B{fila}')
    celda_verificacion = ws_patrimonio.cell(row=fila, column=1)
    celda_verificacion.value = "VERIFICACIÓN: ACTIVO = PASIVO + PATRIMONIO"
    celda_verificacion.font = Font(bold=True, italic=True, color=COLOR_ERROR)
    celda_verificacion.alignment = Alignment(horizontal="center")
    
    # Aplicar bordes
    #aplicar_bordes_tabla(ws_patrimonio, 3, 1, fila-2, 3)
    df_patrimonio = pd.DataFrame({
        'PATRIMONIO NETO': [
            'Capital social',
            'Prima de emisión',
            'Reserva legal',
            'Otras reservas',
            'Resultados ejercicios anteriores',
            'Resultado del ejercicio',
            'Ajustes por cambio de valor',
            'Subvenciones de capital'
        ],
        f'Año {año_actual-1}': [0] * 8
    })
    df_patrimonio.to_excel(writer, sheet_name='Balance - Patrimonio', index=False)

    # HOJA 6: DATOS LABORALES
    # ==== HOJA 7: DATOS LABORALES ====
    ws_laboral = wb.create_sheet("Datos Laborales")
    ws_laboral.sheet_properties.tabColor = COLOR_SECUNDARIO
    
    # Título
    ws_laboral.merge_cells('A1:D1')
    titulo_laboral = ws_laboral['A1']
    titulo_laboral.value = "DATOS LABORALES Y REESTRUCTURACIÓN"
    aplicar_estilo(titulo_laboral, ESTILO_HEADER)
    ws_laboral.row_dimensions[1].height = 25
    
    # Headers
    headers_laboral = ['Campo', 'Valor', 'Unidad', 'Notas']
    for col, header in enumerate(headers_laboral, 1):
        celda = ws_laboral.cell(row=3, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    
    # Ancho de columnas
    ws_laboral.column_dimensions['A'].width = 30
    ws_laboral.column_dimensions['B'].width = 15
    ws_laboral.column_dimensions['C'].width = 12
    ws_laboral.column_dimensions['D'].width = 35
    
    # Datos laborales
    fila = 4
    datos_laborales = [
        ('Número de empleados', 10, 'personas', 'Total plantilla actual'),
        ('Coste medio por empleado', 35000, '€/año', 'Incluye SS empresa'),
        ('Antigüedad media plantilla', 5, 'años', 'Media años en empresa'),
        ('Rotación anual esperada', 10, '%', 'Tasa de rotación'),
        ('', '', '', ''),  # Línea vacía
        ('REESTRUCTURACIÓN', '', '', ''),
        ('¿Reestructuración prevista?', 'No', 'Sí/No', 'Despidos previstos'),
        ('% plantilla afectada', 0, '%', 'Solo si reestructuración'),
        ('Días indemnización por año', 20, 'días', '20/33/45 o personalizado')
    ]
    
    for campo, valor, unidad, nota in datos_laborales:
        if campo == '':
            fila += 1
            continue
        elif campo == 'REESTRUCTURACIÓN':
            celda = ws_laboral.cell(row=fila, column=1)
            celda.value = campo
            celda.font = Font(bold=True, size=11, color="FFFFFF")
            celda.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
            ws_laboral.merge_cells(f'A{fila}:D{fila}')
        else:
            ws_laboral.cell(row=fila, column=1).value = campo
            celda_valor = ws_laboral.cell(row=fila, column=2)
            celda_valor.value = valor
            if isinstance(valor, (int, float)) and unidad != 'Sí/No':
                if unidad == '%':
                    celda_valor.value = valor / 100  # Convertir a decimal
                    celda_valor.number_format = '0.00%'
                else:
                    celda_valor.number_format = '#,##0'
            celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
            ws_laboral.cell(row=fila, column=3).value = unidad
            ws_laboral.cell(row=fila, column=4).value = nota
            ws_laboral.cell(row=fila, column=4).font = Font(italic=True, size=9, color="6B7280")
        fila += 1
    
    # Aplicar bordes
    #aplicar_bordes_tabla(ws_laboral, 3, 1, fila-1, 4)
    df_laboral = pd.DataFrame({
        'Campo': [
            'Número de empleados',
            'Coste medio por empleado',
            'Antigüedad media plantilla (años)',
            'Rotación anual esperada (%)',
            '¿Reestructuración prevista?',
            '% plantilla afectada',
            'Días indemnización por año'
        ],
        'Valor': [10, 35000, 5.0, 10.0, 'No', 0, 20],
        'Notas': [
            'Total empleados',
            'Incluye salario + SS + beneficios',
            'Años promedio',
            'Porcentaje anual',
            'Sí/No',
            'Solo si reestructuración = Sí',
            '20/33/45 días o personalizado'
        ]
    })
    df_laboral.to_excel(writer, sheet_name='Datos Laborales', index=False)
    
    # HOJA 7: FINANCIACIÓN
    # ==== HOJA 8: LÍNEAS DE FINANCIACIÓN ====
    ws_financiacion = wb.create_sheet("Líneas Financiación")
    ws_financiacion.sheet_properties.tabColor = COLOR_SECUNDARIO
    
    # Título
    ws_financiacion.merge_cells('A1:F1')
    titulo_financ = ws_financiacion['A1']
    titulo_financ.value = "LÍNEAS DE FINANCIACIÓN BANCARIA"
    aplicar_estilo(titulo_financ, ESTILO_HEADER)
    ws_financiacion.row_dimensions[1].height = 25
    
    # Headers
    headers_financ = ['Tipo', 'Banco', 'Límite (€)', 'Dispuesto (€)', 'Disponible (€)', 'Tipo interés (%)']
    for col, header in enumerate(headers_financ, 1):
        celda = ws_financiacion.cell(row=3, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    
    # Ancho de columnas
    ws_financiacion.column_dimensions['A'].width = 20
    ws_financiacion.column_dimensions['B'].width = 25
    ws_financiacion.column_dimensions['C'].width = 15
    ws_financiacion.column_dimensions['D'].width = 15
    ws_financiacion.column_dimensions['E'].width = 15
    ws_financiacion.column_dimensions['F'].width = 12
    
    # Líneas de ejemplo
    fila = 4
    lineas_ejemplo = [
        ('Póliza crédito', 'Banco principal', 500000, 250000, '=C4-D4', 4.5),
        ('Confirming', '', 0, 0, '=C5-D5', 0),
        ('Descuento pagarés', '', 0, 0, '=C6-D6', 0),
        ('Línea avales', '', 0, 0, '=C7-D7', 0),
        ('', '', '', '', '', ''),
        ('TOTAL LÍNEAS', '', 0, 0, 0, 0)
    ]
    
    for tipo, banco, limite, dispuesto, disponible, tipo_int in lineas_ejemplo:
        if tipo == '':
            fila += 1
            continue
        elif tipo == 'TOTAL LÍNEAS':
            # Fila de totales
            celda_total = ws_financiacion.cell(row=fila, column=1)
            celda_total.value = tipo
            aplicar_estilo(celda_total, ESTILO_TOTAL)
            
            #for col, formula in enumerate([limite, dispuesto, disponible, tipo_int], 3):
                #celda = ws_financiacion.cell(row=fila, column=col)
                #celda.value = formula
                #aplicar_estilo(celda, ESTILO_TOTAL)
                #if col < 6:
                    #celda.number_format = '#,##0'
                #else:
                    #celda.number_format = '0.00%'
        else:
            # Datos normales
            ws_financiacion.cell(row=fila, column=1).value = tipo
            ws_financiacion.cell(row=fila, column=2).value = banco
            
            celda_limite = ws_financiacion.cell(row=fila, column=3)
            celda_limite.value = limite
            celda_limite.number_format = '#,##0'
            celda_limite.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
            
            celda_dispuesto = ws_financiacion.cell(row=fila, column=4)
            celda_dispuesto.value = dispuesto
            celda_dispuesto.number_format = '#,##0'
            celda_dispuesto.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
            
            celda_disponible = ws_financiacion.cell(row=fila, column=5)
            celda_disponible.value = disponible
            celda_disponible.number_format = '#,##0'
            celda_disponible.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            
            celda_tipo = ws_financiacion.cell(row=fila, column=6)
            celda_tipo.value = tipo_int / 100  # Convertir a decimal para porcentaje
            celda_tipo.number_format = '0.00%'
            celda_tipo.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        
        fila += 1
    
    # Aplicar bordes
    #aplicar_bordes_tabla(ws_financiacion, 3, 1, fila-1, 6)
    df_financiacion = pd.DataFrame({
        'Tipo': ['Póliza crédito', '', ''],
        'Banco': ['Banco principal', '', ''],
        'Límite': [500000, 0, 0],
        'Dispuesto': [250000, 0, 0],
        'Tipo interés (%)': [4.5, 0, 0]
    })
    df_financiacion.to_excel(writer, sheet_name='Líneas Financiación', index=False)
    
    # HOJA 8: PROYECCIONES Y PARÁMETROS
    # ==== HOJA 9: PROYECCIONES Y PARÁMETROS ====
    ws_proyecciones = wb.create_sheet("Proyecciones y Parámetros")
    ws_proyecciones.sheet_properties.tabColor = COLOR_SECUNDARIO
    
    # Título
    ws_proyecciones.merge_cells('A1:D1')
    titulo_proy = ws_proyecciones['A1']
    titulo_proy.value = "PROYECCIONES Y PARÁMETROS OPERATIVOS"
    aplicar_estilo(titulo_proy, ESTILO_HEADER)
    ws_proyecciones.row_dimensions[1].height = 25
    
    # SECCIÓN 1: PLAN DE INVERSIONES
    fila = 3
    # Título sección
    celda_inv = ws_proyecciones.cell(row=fila, column=1)
    celda_inv.value = "PLAN DE INVERSIONES (CAPEX)"
    celda_inv.font = Font(bold=True, size=11, color="FFFFFF")
    celda_inv.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    ws_proyecciones.merge_cells(f'A{fila}:D{fila}')
    fila += 1
    
    # Headers inversiones
    headers_inv = ['Concepto', 'Importe (€)', 'Años', 'Notas']
    for col, header in enumerate(headers_inv, 1):
        celda = ws_proyecciones.cell(row=fila, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    fila += 1
    
    # Datos inversiones
    for i in range(1, 6):
        ws_proyecciones.cell(row=fila, column=1).value = f'Inversión Año {i}'
        celda_valor = ws_proyecciones.cell(row=fila, column=2)
        celda_valor.value = 0
        celda_valor.number_format = '#,##0'
        celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        ws_proyecciones.cell(row=fila, column=4).value = 'Maquinaria, equipos, software...'
        ws_proyecciones.cell(row=fila, column=4).font = Font(italic=True, size=9, color="6B7280")
        fila += 1
    
    # Vida útil
    ws_proyecciones.cell(row=fila, column=1).value = 'Vida útil media'
    celda_vida = ws_proyecciones.cell(row=fila, column=2)
    celda_vida.value = 10
    celda_vida.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
    ws_proyecciones.cell(row=fila, column=3).value = 'años'
    ws_proyecciones.cell(row=fila, column=4).value = 'Para cálculo amortización'
    ws_proyecciones.cell(row=fila, column=4).font = Font(italic=True, size=9, color="6B7280")
    fila += 2
    
    # SECCIÓN 2: PARÁMETROS OPERATIVOS
    celda_param = ws_proyecciones.cell(row=fila, column=1)
    celda_param.value = "PARÁMETROS OPERATIVOS"
    celda_param.font = Font(bold=True, size=11, color="FFFFFF")
    celda_param.fill = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    ws_proyecciones.merge_cells(f'A{fila}:D{fila}')
    fila += 1
    
    # Headers parámetros
    for col, header in enumerate(['Parámetro', 'Valor', 'Unidad', 'Impacto'], 1):
        celda = ws_proyecciones.cell(row=fila, column=col)
        celda.value = header
        aplicar_estilo(celda, ESTILO_SUBTITULO)
    fila += 1
    
    # Datos parámetros
    parametros = [
        ('Días de cobro', 60, 'días', 'Afecta a clientes'),
        ('Días de pago', 30, 'días', 'Afecta a proveedores'),
        ('Días de stock', 45, 'días', 'Afecta a inventario'),
        ('Crecimiento extraordinario', 0, '%', 'Eventos especiales'),
        ('Tasa impositiva', 25, '%', 'Impuesto sociedades'),
        ('Dividendos sobre beneficio', 0, '%', 'Política dividendos')
    ]
    
    for param, valor, unidad, impacto in parametros:
        ws_proyecciones.cell(row=fila, column=1).value = param
        celda_valor = ws_proyecciones.cell(row=fila, column=2)
        celda_valor.value = valor
        celda_valor.number_format = '0' if unidad == '%' else '#,##0'
        celda_valor.fill = PatternFill(start_color="FFFEF0", end_color="FFFEF0", fill_type="solid")
        ws_proyecciones.cell(row=fila, column=3).value = unidad
        ws_proyecciones.cell(row=fila, column=4).value = impacto
        ws_proyecciones.cell(row=fila, column=4).font = Font(italic=True, size=9, color="6B7280")
        fila += 1
    
    # Aplicar bordes a todo
    #aplicar_bordes_tabla(ws_proyecciones, 4, 1, 10, 4)  # Inversiones
    #aplicar_bordes_tabla(ws_proyecciones, 14, 1, 20, 4)  # Parámetros
    
    # Ancho de columnas
    ws_proyecciones.column_dimensions['A'].width = 25
    ws_proyecciones.column_dimensions['B'].width = 15
    ws_proyecciones.column_dimensions['C'].width = 10
    ws_proyecciones.column_dimensions['D'].width = 30
    df_proyecciones = pd.DataFrame({
        'Concepto': [
            'PLAN DE INVERSIONES',
            'Inversión Año 1',
            'Inversión Año 2',
            'Inversión Año 3',
            'Inversión Año 4',
            'Inversión Año 5',
            'Vida útil media (años)',
            '',
            'PARÁMETROS OPERATIVOS',
            'Días de cobro',
            'Días de pago',
            'Días de stock',
            'Eventos extraordinarios (%)'
        ],
        'Valor': [
            '',
            0,
            0,
            0,
            0,
            0,
            10,
            '',
            '',
            60,
            30,
            45,
            0
        ]
    })
    df_proyecciones.to_excel(writer, sheet_name='Proyecciones y Parámetros', index=False)

    # ==== GUARDAR CON TODOS LOS DATAFRAMES RESTANTES ====
    # Guardar el workbook de openpyxl primero
    # ==== AÑADIR NOTAS CON OPCIONES VÁLIDAS ====
    # Como Excel Mac no muestra las validaciones, añadimos las opciones en comentarios
    ws_info['C5'].value = "Opciones: General, Hostelería, Tecnología, Ecommerce, Industrial"
    ws_info['C8'].value = "Opciones: Sí, No"
    ws_info['C9'].value = "Opciones: Sí, No"
    ws_info['C10'].value = "Opciones: EUR, USD, GBP, CHF"
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# Funciones auxiliares globales
def safe_int(value, default=0):
    try:
        if pd.isna(value) or value == '':
            return default
        return int(float(value))
    except:
        return default

def safe_float(value, default=0.0):
    try:
        if pd.isna(value) or value == '':
            return default
        return float(value)
    except:
        return default

def leer_excel_datos(uploaded_file):
    """
    Lee un archivo Excel con el formato de la plantilla y extrae todos los datos
    """
    try:
        # Leer todas las hojas del Excel
        excel_file = pd.ExcelFile(uploaded_file)
        print(f"Hojas encontradas en el archivo: {excel_file.sheet_names}")
        
        # Verificar que existen las hojas esperadas
        hojas_requeridas = ['Informacion General', 'Datos Históricos PYL', 'Balance - Activo',
                           'Balance - Pasivo', 'Balance - Patrimonio', 'Datos Laborales',
                           'Líneas Financiación', 'Proyecciones y Parámetros']
        
        for hoja in hojas_requeridas:
            if hoja not in excel_file.sheet_names:
                st.error(f"Falta la hoja: '{hoja}'")
                return None
        
        datos = {}
       
        # LEER INFORMACIÓN GENERAL
        df_info = pd.read_excel(uploaded_file, sheet_name='Informacion General')
        info_dict = dict(zip(df_info['Campo'], df_info['Valor']))
        
        datos['info_general'] = {
            'nombre_empresa': str(info_dict.get('Nombre de la empresa', '')),
            'sector': str(info_dict.get('Sector', 'Otro')),
            'pais': str(info_dict.get('País', 'España')),
            'año_fundacion': safe_int(info_dict.get('Año de Fundación', datetime.now().year - 5)) if pd.notna(info_dict.get('Año de Fundación')) else datetime.now().year - 5,
            'empresa_familiar': info_dict.get('¿Empresa familiar?', 'No'),
            'empresa_auditada': info_dict.get('¿Cuentas auditadas?', 'Sí'),
            'moneda': str(info_dict.get('Moneda', 'EUR'))
        }
        
        # LEER DATOS HISTÓRICOS P&L
        df_pyl = pd.read_excel(uploaded_file, sheet_name='Datos Históricos PYL')
        pyl_dict = df_pyl.set_index('Concepto').to_dict()
        
        # Obtener los años de las columnas
        años_pyl = [col for col in df_pyl.columns if col != 'Concepto']
        
        datos['pyl_historico'] = {
            'años': años_pyl,
            'ventas': [pyl_dict[año].get('Ventas', 0) for año in años_pyl],
            'costos_variables_pct': pyl_dict[años_pyl[-1]].get('Costos Variables (%)', 40),
            'gastos_personal': pyl_dict[años_pyl[-1]].get('Gastos de Personal', 0),
            'gastos_generales': pyl_dict[años_pyl[-1]].get('Gastos Generales', 0),
            'gastos_marketing': pyl_dict[años_pyl[-1]].get('Gastos de Marketing', 0)
        }
        # DEBUG: Imprimir datos leídos
        print("\n=== DATOS P&L LEÍDOS DEL EXCEL ===")
        print(f"Ventas último año: {datos['pyl_historico']['ventas'][-1]:,.0f}")
        print(f"Costos variables: {datos['pyl_historico']['costos_variables_pct']}%")
        print(f"Gastos personal: {datos['pyl_historico']['gastos_personal']:,.0f}")
        print(f"Gastos generales: {datos['pyl_historico']['gastos_generales']:,.0f}")
        print(f"Gastos marketing: {datos['pyl_historico']['gastos_marketing']:,.0f}")
        print("=====================================\n")

        # LEER BALANCE - ACTIVO
        df_activo = pd.read_excel(uploaded_file, sheet_name='Balance - Activo')
        activo_dict = dict(zip(df_activo.iloc[:, 0], df_activo.iloc[:, 1]))
        
        datos['balance_activo'] = {
            'tesoreria_inicial': safe_float(activo_dict.get('Caja y bancos', 0)),
            'inversiones_cp': safe_float(activo_dict.get('Inversiones financieras temporales', 0)),
            'clientes_inicial': safe_float(activo_dict.get('Clientes comerciales', 0)),
            'otros_deudores': safe_float(activo_dict.get('Otros deudores', 0)),
            'admin_publica_deudora': safe_float(activo_dict.get('Administraciones públicas deudoras', 0)),
            'inventario_inicial': safe_float(activo_dict.get('Inventarios', 0)),
            'gastos_anticipados': safe_float(activo_dict.get('Gastos anticipados', 0)),
            'activos_impuesto_diferido_cp': safe_float(activo_dict.get('Activos por impuesto diferido CP', 0)),
            'activo_fijo_bruto': safe_float(activo_dict.get('Inmovilizado material bruto', 0)),
            'depreciacion_acumulada': safe_float(activo_dict.get('Amortización acumulada material', 0)),
            'activos_intangibles': safe_float(activo_dict.get('Activos intangibles brutos', 0)),
            'amortizacion_intangibles': safe_float(activo_dict.get('Amortización acumulada intangibles', 0)),
            'inversiones_lp': safe_float(activo_dict.get('Participaciones en empresas', 0)),
            'creditos_lp': safe_float(activo_dict.get('Créditos a largo plazo', 0)),
            'fianzas_depositos': safe_float(activo_dict.get('Fianzas y depósitos', 0)),
            'activos_impuesto_diferido_lp': safe_float(activo_dict.get('Activos por impuesto diferido LP', 0))
        }

        # LEER BALANCE - PASIVO
        df_pasivo = pd.read_excel(uploaded_file, sheet_name='Balance - Pasivo')
        pasivo_dict = dict(zip(df_pasivo.iloc[:, 0], df_pasivo.iloc[:, 1]))
        
        datos['balance_pasivo'] = {
            'proveedores_inicial': safe_float(pasivo_dict.get('Proveedores comerciales', 0)),
            'acreedores_servicios': safe_float(pasivo_dict.get('Acreedores por servicios', 0)),
            'anticipos_clientes': safe_float(pasivo_dict.get('Anticipos de clientes', 0)),
            'remuneraciones_pendientes': safe_float(pasivo_dict.get('Remuneraciones pendientes', 0)),
            'admin_publica_acreedora': safe_float(pasivo_dict.get('Administraciones públicas acreedoras', 0)),
            'provisiones_cp': safe_float(pasivo_dict.get('Provisiones a corto plazo', 0)),
            'otros_pasivos_cp': safe_float(pasivo_dict.get('Otros pasivos corrientes', 0)),
            'prestamo_principal': safe_float(pasivo_dict.get('Préstamos bancarios LP', 0)),
            'hipoteca_importe_original': safe_float(pasivo_dict.get('Hipoteca importe original', 0)),
            'hipoteca_meses_transcurridos': safe_int(pasivo_dict.get('Hipoteca meses transcurridos', 0)),
            'leasing_total': safe_float(pasivo_dict.get('Leasing pendiente', 0)),
            'otros_prestamos_lp': safe_float(pasivo_dict.get('Otros préstamos LP', 0)),
            'provisiones_riesgos': safe_float(pasivo_dict.get('Provisiones para riesgos', 0)),
            'otras_provisiones_lp': safe_float(pasivo_dict.get('Otras provisiones LP', 0)),
            'pasivos_impuesto_diferido': safe_float(pasivo_dict.get('Pasivos por impuesto diferido', 0))
        }

        # LEER BALANCE - PATRIMONIO
        df_patrimonio = pd.read_excel(uploaded_file, sheet_name='Balance - Patrimonio')
        patrimonio_dict = dict(zip(df_patrimonio.iloc[:, 0], df_patrimonio.iloc[:, 1]))
        
        datos['balance_patrimonio'] = {
            'capital_social': safe_float(patrimonio_dict.get('Capital social', 100000)),
            'prima_emision': safe_float(patrimonio_dict.get('Prima de emisión', 0)),
            'reserva_legal': safe_float(patrimonio_dict.get('Reserva legal', 20000)),
            'reservas': safe_float(patrimonio_dict.get('Otras reservas', 0)),
            'resultados_acumulados': safe_float(patrimonio_dict.get('Resultados ejercicios anteriores', 0)),
            'resultado_ejercicio': safe_float(patrimonio_dict.get('Resultado del ejercicio', 0)),
            'ajustes_valor': safe_float(patrimonio_dict.get('Ajustes por cambio de valor', 0)),
            'subvenciones': safe_float(patrimonio_dict.get('Subvenciones de capital', 0))
        }
        
        # LEER DATOS LABORALES
        df_laboral = pd.read_excel(uploaded_file, sheet_name='Datos Laborales')
        laboral_dict = dict(zip(df_laboral['Concepto'], df_laboral['Valor']))
        
        datos['datos_laborales'] = {
            'num_empleados': safe_int(laboral_dict.get('Número de empleados', 10)) if pd.notna(laboral_dict.get('Número de empleados')) else 10,
            'coste_medio_empleado': safe_float(laboral_dict.get('Coste medio por empleado', 35000)),
            'antiguedad_media': safe_float(laboral_dict.get('Antigüedad media plantilla (años)', 5.0)),
            'rotacion_anual': safe_float(laboral_dict.get('Rotación anual esperada (%)', 10.0)),
            'reestructuracion_prevista': laboral_dict.get('¿Reestructuración prevista?', 'No') == 'Sí',
            'porcentaje_afectados': safe_float(laboral_dict.get('% plantilla afectada', 0)),
            'dias_indemnizacion': safe_int(laboral_dict.get('Días indemnización por año', 20))
        }
# LEER LÍNEAS DE FINANCIACIÓN
        df_financiacion = pd.read_excel(uploaded_file, sheet_name='Líneas Financiación')
        lineas_financiacion = []
        for _, row in df_financiacion.iterrows():
            if pd.notna(row.get('Tipo', '')) and row.get('Tipo', '') != '':
                lineas_financiacion.append({
                    'tipo': row.get('Tipo', 'Póliza crédito'),
                    'banco': row.get('Banco', 'Banco principal'),
                    'limite': safe_float(row.get('Límite', 0)),
                    'dispuesto': safe_float(row.get('Dispuesto', 0)),
                    'tipo_interes': safe_float(row.get('Tipo interés (%)', 4.5))
                })
        
        datos['lineas_financiacion'] = lineas_financiacion
        
        # LEER PROYECCIONES Y PARÁMETROS
        df_proyecciones = pd.read_excel(uploaded_file, sheet_name='Proyecciones y Parámetros')
        proyecciones_dict = dict(zip(df_proyecciones['Concepto'], df_proyecciones['Valor']))
        
        datos['proyecciones'] = {
            'capex_año1': safe_float(proyecciones_dict.get('Inversión Año 1', 0)),
            'capex_año2': safe_float(proyecciones_dict.get('Inversión Año 2', 0)),
            'capex_año3': safe_float(proyecciones_dict.get('Inversión Año 3', 0)),
            'capex_año4': safe_float(proyecciones_dict.get('Inversión Año 4', 0)),
            'capex_año5': safe_float(proyecciones_dict.get('Inversión Año 5', 0)),
            'vida_util': safe_int(proyecciones_dict.get('Vida útil media (años)', 10)),
            'dias_cobro': safe_int(proyecciones_dict.get('Días de cobro', 60)),
            'dias_pago': safe_int(proyecciones_dict.get('Días de pago', 30)),
            'dias_stock': safe_int(proyecciones_dict.get('Días de stock', 45)),
            'crecimiento_extraordinario': safe_float(proyecciones_dict.get('Eventos extraordinarios (%)', 0))
        }
        
        return datos
        
    except Exception as e:
        st.error(f"Error al leer el archivo Excel: {str(e)}")
        return None



